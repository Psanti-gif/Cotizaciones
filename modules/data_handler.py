from openpyxl import Workbook, load_workbook
from fpdf import FPDF
from tkinter import messagebox

EXCEL_FILE = "cotizaciones.xlsx"

def nueva_cotizacion():
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Cotización"
        ws.append(["Descripción", "Cantidad", "Costo con IVA", "Subtotal", "Total"])
        wb.save(EXCEL_FILE)
        messagebox.showinfo("Éxito", "Nueva cotización iniciada.")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo iniciar la cotización: {e}")

def agregar_producto(descripcion, cantidad, costo):
    if not descripcion or not cantidad or not costo:
        messagebox.showwarning("Campos vacíos", "Todos los campos son obligatorios.")
        return

    try:
        cantidad = int(cantidad)
        costo = float(costo)
        subtotal = cantidad * costo
        total = subtotal
        
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append([descripcion, cantidad, costo, subtotal, total])
        wb.save(EXCEL_FILE)
        messagebox.showinfo("Éxito", "Producto agregado a la cotización.")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo agregar el producto: {e}")

def generar_pdf():
    try:
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        pdf.cell(200, 10, txt="Cotización", ln=True, align='C')
        
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            pdf.cell(0, 10, txt=" | ".join(map(str, row)), ln=True)
        
        pdf_file = "cotizacion.pdf"
        pdf.output(pdf_file)
        messagebox.showinfo("Éxito", f"PDF generado: {pdf_file}")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo generar el PDF: {e}")
