import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
from fpdf import FPDF


EXCEL_FILE = "cotizaciones.xlsx"


def nueva_cotizacion():
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Cotización"
        ws.append(["Descripción", "Cantidad", "Costo con IVA", "Subtotal", "Total"])
        wb.save(EXCEL_FILE)
        messagebox.showinfo("Éxito", "Nueva cotización iniciada.")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo iniciar la cotización: {e}")


def agregar_producto():
    descripcion = descripcion_var.get()
    cantidad = cantidad_var.get()
    costo = costo_var.get()
    
    if not descripcion or not cantidad or not costo:
        messagebox.showwarning("Campos vacíos", "Todos los campos son obligatorios.")
        return

    try:
        cantidad = int(cantidad)
        costo = float(costo)
        subtotal = cantidad * costo
        total = subtotal  
        
        
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append([descripcion, cantidad, costo, subtotal, total])
        wb.save(EXCEL_FILE)

        # Limpiar campos
        descripcion_var.set("")
        cantidad_var.set("")
        costo_var.set("")

        messagebox.showinfo("Éxito", "Producto agregado a la cotización.")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo agregar el producto: {e}")


def generar_pdf():
    try:
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        pdf.cell(200, 10, txt="Cotización", ln=True, align='C')
        
        
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            pdf.cell(0, 10, txt=" | ".join(map(str, row)), ln=True)
        
        
        pdf_file = "cotizacion.pdf"
        pdf.output(pdf_file)
        messagebox.showinfo("Éxito", f"PDF generado: {pdf_file}")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo generar el PDF: {e}")


app = tk.Tk()
app.title("Gestor de Cotizaciones")


descripcion_var = tk.StringVar()
cantidad_var = tk.StringVar()
costo_var = tk.StringVar()


tk.Label(app, text="Descripción:").grid(row=0, column=0, padx=10, pady=5)
tk.Entry(app, textvariable=descripcion_var).grid(row=0, column=1, padx=10, pady=5)

tk.Label(app, text="Cantidad:").grid(row=1, column=0, padx=10, pady=5)
tk.Entry(app, textvariable=cantidad_var).grid(row=1, column=1, padx=10, pady=5)

tk.Label(app, text="Costo con IVA:").grid(row=2, column=0, padx=10, pady=5)
tk.Entry(app, textvariable=costo_var).grid(row=2, column=1, padx=10, pady=5)

tk.Button(app, text="Agregar Producto", command=agregar_producto).grid(row=3, column=0, columnspan=2, pady=10)
tk.Button(app, text="Nueva Cotización", command=nueva_cotizacion).grid(row=4, column=0, pady=10)
tk.Button(app, text="Generar PDF", command=generar_pdf).grid(row=4, column=1, pady=10)

app.mainloop()
